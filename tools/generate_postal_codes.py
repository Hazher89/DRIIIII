#!/usr/bin/env python3
"""Generer assets/data/norway_postal_codes.json fra POSTKODE.xlsx i prosjektrot."""
import json
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "POSTKODE.xlsx"
OUT = ROOT / "assets" / "data" / "norway_postal_codes.json"


def main() -> None:
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    sheet = "areas_2022-11-21_18-06-41.csv"
    if sheet not in wb.sheetnames:
        sheet = wb.sheetnames[-1]
    ws = wb[sheet]
    by_pc: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        pc = str(row[0]).strip().zfill(4)
        if not re.match(r"^\d{4}$", pc):
            continue
        sted = str(row[1] or "").strip()
        if sted:
            by_pc[pc] = sted
    wb.close()

    OUT.parent.mkdir(parents=True, exist_ok=True)
    payload = {"version": 1, "count": len(by_pc), "postal_codes": by_pc}
    OUT.write_text(json.dumps(payload, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    print(f"Wrote {len(by_pc)} postnummer → {OUT}")


if __name__ == "__main__":
    main()
